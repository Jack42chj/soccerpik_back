from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import json
import pandas as pd
import joblib

app = Flask(__name__)
CORS(app, resources={r'/*': {'origins': '*'}}, supports_credentials=True)

@app.route("/", methods=['GET'])
def router():
    print("root url!")
    return jsonify("root url")

@app.route("/search", methods=['POST'])
def search():
    data = request.get_json()
    length = data['length']
    width = data['width']
    height = data['height']
    field = data['field']

    data = pd.read_excel('soccer.xlsx')
    data.columns = ['Name', 'Width', 'Height', 'Stud']

    cluster_model = joblib.load('cluster_model')
    dtree_model = joblib.load('dtree_model')

    err = 1
    while(err == 1):
        foot_size = float(length)
        foot_Width = float(width)
        foot_Height = float(height)
        Stud = field
        if Stud == 'FG':
            Stud = 0
            err = 0
        elif Stud == 'AG':
            Stud = 1
            err = 0
        elif Stud == 'SG':
            Stud = 2
            err = 0
        elif Stud == 'TF':
            Stud = 3
            err = 0
        elif Stud == 'IC':
            Stud = 4
            err = 0
        else:
            print('please enter it again')

    Stud = Stud * 1000000
    Width = (foot_Width / foot_size) * 10000000
    Height = (foot_Height / foot_size) * 10000000
    user = [[Width, Height, Stud]]

    user_cluster = int(dtree_model.predict(user))

    result = []
    for index in data.index:
        if cluster_model.labels_[index] == user_cluster:
            result.append(data.loc[index,'Name'])

    result_excel = pd.DataFrame(result)
    result_excel.columns = ['Name']

    result_excel.to_excel('result.xlsx')
    return jsonify({'status': True})

@app.route("/result", methods=['GET'])
def result():
    xlsxname = openpyxl.load_workbook('result.xlsx', read_only=True)
    sheet = xlsxname.worksheets[0]

    data = {}
    data['result'] = []

    key_list = []
    for i in range(1, sheet.max_column + 1):
        key_list.append(sheet.cell(row=1,column=i).value)

    data_list = {}
    for i in range(2, sheet.max_row + 1):
        temp_dict = {}
        for j in range(1, sheet.max_column + 1):
            val = sheet.cell(row=i, column=j).value
            temp_dict[key_list[j - 1]] = val
        data['result'].append(temp_dict)
    
    xlsxname.close()
    result = json.dumps(data)
    return result

if __name__ == "__main__":
    app.run(host="0.0.0.0", port="5000")